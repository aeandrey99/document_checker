#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import datetime
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

class ReportManager:
    """
    Класс для управления отчетами проверки документов
    """
    def __init__(self, app_instance):
        self.app = app_instance
        
        # Для хранения существующих результатов
        self.existing_data = []
        
        # Номер текущего запуска
        self.current_run_id = 0
    
    def load_existing_results(self):
        """Загружает существующие данные из отчета"""
        # Получаем путь к файлу отчета
        report_path = self.app.output_path.get()
        
        # Проверяем наличие файла
        if os.path.exists(report_path):
            try:
                # Загрузка существующих данных из Excel
                df = pd.read_excel(report_path)
                self.existing_data = df.to_dict('records')
                
                # Определяем максимальный порядковый номер запуска
                if not self.existing_data:
                    self.current_run_id = 0  # При первом запуске будет увеличено до 1
                else:
                    # Проверяем наличие столбца "Порядковый номер запуска"
                    if 'Порядковый номер запуска' in df.columns:
                        self.current_run_id = df['Порядковый номер запуска'].max()
                    else:
                        # Если столбца нет, добавляем его с значением 0 для существующих записей
                        for row in self.existing_data:
                            row['Порядковый номер запуска'] = 0
                        self.current_run_id = 0
                
                # Обновляем отображение текущего номера запуска в интерфейсе
                if hasattr(self.app, 'run_id_label'):
                    self.app.run_id_label.config(text=str(self.current_run_id))
                    
            except Exception as e:
                self.app.show_warning("Предупреждение", f"Не удалось загрузить существующий отчет: {str(e)}")
                self.existing_data = []
                self.current_run_id = 0
        else:
            self.existing_data = []
            self.current_run_id = 0
        
        # Обновляем отображение текущего номера запуска в интерфейсе
        if hasattr(self.app, 'run_id_label'):
            self.app.run_id_label.config(text=str(self.current_run_id))
    
    def load_run_id_from_report(self):
        """Загружает номер запуска из файла отчета"""
        report_path = self.app.output_path.get()
        
        # Проверяем наличие файла
        if os.path.exists(report_path):
            try:
                # Загрузка данных из отчета
                df = pd.read_excel(report_path)
                
                # Проверяем наличие столбца с номером запуска
                if 'Порядковый номер запуска' in df.columns and not df.empty:
                    # Определяем максимальный номер запуска и увеличиваем на 1
                    max_run_id = df['Порядковый номер запуска'].max()
                    self.current_run_id = max_run_id
                else:
                    # Если столбца нет или файл пустой, начинаем с 1
                    self.current_run_id = 0
                    
                # Обновляем отображение номера запуска
                self.app.run_id_label.config(text=str(self.current_run_id))
                
            except Exception as e:
                self.app.show_warning("Предупреждение", f"Не удалось прочитать номер запуска из отчета: {str(e)}")
                self.current_run_id = 0
        else:
            # Если файла нет, начинаем с 0 (первый запуск будет 1)
            self.current_run_id = 0
            self.app.run_id_label.config(text=str(self.current_run_id))
    
    def save_results_to_excel(self, results):
        """
        Сохраняет результаты проверки в Excel
        
        Args:
            results (list): Список результатов проверки
        """
        # Получаем путь к файлу отчета
        report_path = self.app.output_path.get()
        
        # Проверяем, указан ли путь
        if not report_path:
            # Если путь не указан, используем директорию приложения
            if getattr(sys, 'frozen', False):
                # Скомпилированный вариант
                app_dir = os.path.dirname(sys.executable)
            else:
                # Запуск из Python скрипта
                app_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            
            # Формируем путь к файлу отчета
            report_path = os.path.join(app_dir, "Отчет_проверки_документов.xlsx")
            self.app.output_path.set(report_path)
        
        # Нормализуем путь к отчету
        from app.core.file_utils import normalize_path
        report_path = normalize_path(report_path)
        
        # Текущая дата и время
        now = datetime.datetime.now()
        check_date = now.strftime("%d.%m.%Y %H:%M:%S")
        
        # Сначала проверяем, существует ли файл и нужно ли загрузить его заново
        # (это важно, если файл был изменен другой программой)
        if os.path.exists(report_path):
            try:
                df_existing = pd.read_excel(report_path)
                self.existing_data = df_existing.to_dict('records')
            except Exception as e:
                # Если ошибка при чтении файла, продолжаем с текущими данными
                print(f"Предупреждение: Ошибка чтения отчета: {str(e)}")
                # Если existing_data пустой, пытаемся создать новый файл
                if not self.existing_data:
                    self.existing_data = []
        
        # Подготовка данных нового запуска
        new_data = []
        for result in results:
            # Нормализуем путь к файлу
            file_path = normalize_path(result['file_path'])
            
            new_data.append({
                'Порядковый номер': len(self.existing_data) + len(new_data) + 1,
                'Порядковый номер запуска': self.current_run_id,  # Используем текущий номер запуска
                'Дата проверки': check_date,
                'Имя файла': result['file_name'],
                'Тип файла': result['file_type'],
                'Путь к файлу': file_path,
                'Результат проверки': result['result'],
                'Комментарий по результатам проверки': result['comment']
            })
        
        # Объединение существующих и новых данных
        all_data = self.existing_data + new_data
        
        # Создание DataFrame и сохранение в Excel
        df = pd.DataFrame(all_data)
        
        # Проверяем наличие директории
        os.makedirs(os.path.dirname(os.path.abspath(report_path)), exist_ok=True)
        
        # Создание Excel с форматированием
        with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Отчет')
            
            # Применение форматирования
            workbook = writer.book
            worksheet = writer.sheets['Отчет']
            
            # Форматирование заголовков
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
            
            # Цветовое форматирование ячеек результатов
            green_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')  # Светло-зеленый
            red_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')    # Светло-красный
            
            # Создаем гиперссылки на пути к файлам
            path_col_index = df.columns.get_loc('Путь к файлу') + 1  # +1 потому что в openpyxl колонки с 1
            
            for row_num, row_data in enumerate(df.iterrows(), 2):  # Начинаем с 2, т.к. 1 - заголовки
                # Цветовое форматирование ячеек результатов
                result_cell = worksheet.cell(row=row_num, column=df.columns.get_loc('Результат проверки') + 1)
                if result_cell.value == "Пройден":
                    result_cell.fill = green_fill
                elif result_cell.value in ["Не пройден", "Ошибка"]:
                    result_cell.fill = red_fill
                
                # Добавляем гиперссылку к пути файла
                file_path_cell = worksheet.cell(row=row_num, column=path_col_index)
                file_path = file_path_cell.value
                
                if file_path and os.path.exists(file_path):  # Проверяем существование файла
                    file_path_cell.hyperlink = file_path  # Путь уже в формате с обратными слешами
                    file_path_cell.style = "Hyperlink"
            
            # Автоподбор ширины колонок
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Сообщение об успешном сохранении
        self.app.root.after(0, lambda: self.app.show_info("Сохранено", f"Отчет сохранен в:\n{report_path}"))
        
        return report_path