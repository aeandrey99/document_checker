#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import openpyxl
from zipfile import ZipFile
import win32com.client
import pythoncom

def check_excel_file(file_path, file_name, enable_value_search=False, search_values=None):
    """
    Проверка Excel файла с оптимизацией для крупных файлов
    
    Args:
        file_path (str): Путь к файлу
        file_name (str): Имя файла
        enable_value_search (bool): Флаг включения поиска заданных значений
        search_values (list): Список значений для поиска
    """
    try:
        issues = []
        
        if file_path.lower().endswith(('.xlsx', '.xlsm')):
            # Быстрая предварительная проверка для .xlsx и .xlsm без полной загрузки
            try:
                from zipfile import ZipFile
                with ZipFile(file_path) as xlsx_zip:
                    # Проверка на комментарии - смотрим наличие файлов комментариев
                    comment_files = [f for f in xlsx_zip.namelist() if 'comments' in f.lower()]
                    if comment_files:
                        issues.append("комментарии")
                    
                    # Проверка цвета вкладок - загружаем workbook.xml
                    if 'xl/workbook.xml' in xlsx_zip.namelist():
                        workbook_content = xlsx_zip.read('xl/workbook.xml').decode('utf-8', errors='ignore')
                        if 'tabColor' in workbook_content:
                            # Более тщательная проверка на цвета листов
                            if any(color in workbook_content for color in ['FFFF', 'FF0000', 'FFFF00', 'FFF000']):
                                issues.append("цветной лист")
            except Exception as e:
                # В случае ошибки быстрой проверки, продолжаем обычным способом
                pass
                
            # Если проблем с цветами листов не найдено быстрым способом, выполняем более тщательную проверку
            if not "цветной лист" in issues:
                # Используем openpyxl с оптимизированной загрузкой
                options = {
                    'read_only': False,  # Нужно отключить read_only для доступа к цветам
                    'data_only': True    # Без формул
                }
                
                wb = None
                try:
                    wb = openpyxl.load_workbook(file_path, **options)
                    
                    # Проверка на желтые или красные листы
                    for sheet_name in wb.sheetnames:
                        try:
                            sheet = wb[sheet_name]
                            if sheet.sheet_properties and sheet.sheet_properties.tabColor:
                                tab_color = sheet.sheet_properties.tabColor.rgb
                                # Более широкая проверка на различные оттенки желтого и красного
                                if tab_color:
                                    # Различные варианты цветов
                                    yellow_codes = ['FFFF00', 'FFFFF00', 'FFFFFF00', 'FFF0F000', 'FFFFCC']
                                    red_codes = ['FF0000', 'FFFF0000', 'FFF00000', 'FFCC0000']
                                    
                                    # Проверка на желтый/красный цвет
                                    is_yellow = any(code in tab_color.upper() for code in yellow_codes)
                                    is_red = any(code in tab_color.upper() for code in red_codes)
                                    
                                    if is_yellow or is_red:
                                        issues.append(f"цветной лист ({sheet_name})")
                                        break
                        except Exception as e:
                            # Логируем ошибку и продолжаем
                            continue
                    
                    # Если всё еще не найдены цветные листы, проверяем альтернативным способом
                    if not "цветной лист" in issues:
                        for sheet_name in wb.sheetnames:
                            try:
                                sheet = wb[sheet_name]
                                # Дополнительная проверка через свойства листа
                                if hasattr(sheet, '_WorkbookChild__props') and 'tabColor' in sheet._WorkbookChild__props:
                                    issues.append(f"цветной лист ({sheet_name})")
                                    break
                            except:
                                pass
                
                    # Проверка на желтые ячейки - проверяем сектор 500 столбцов на 1500 строк
                    if not "желтые ячейки" in issues:
                        for sheet_name in wb.sheetnames:
                            try:
                                sheet = wb[sheet_name]
                                max_rows = min(1500, sheet.max_row)  # Максимум 1500 строк
                                max_cols = min(500, sheet.max_column)  # Максимум 500 столбцов
                                
                                for row_idx in range(1, max_rows + 1):
                                    for col_idx in range(1, max_cols + 1):
                                        try:
                                            cell = sheet.cell(row=row_idx, column=col_idx)
                                            if cell.fill and hasattr(cell.fill, 'start_color') and hasattr(cell.fill.start_color, 'index'):
                                                fill_color = cell.fill.start_color.index
                                                if isinstance(fill_color, str) and ('FFFF' in fill_color or 'FFFF00' in fill_color or 'FFF0' in fill_color):
                                                    issues.append("желтые ячейки")
                                                    break
                                        except Exception as cell_ex:
                                            # Пропускаем проблемные ячейки
                                            continue
                                    
                                    if "желтые ячейки" in issues:
                                        break
                                        
                                if "желтые ячейки" in issues:
                                    break
                            except Exception:
                                # Если возникла ошибка при проверке листа, продолжаем с следующим
                                continue
                    
                    # Проверка на комментарии - только если еще не найдены
                    if not "комментарии" in issues:
                        for sheet_name in wb.sheetnames:
                            try:
                                sheet = wb[sheet_name]
                                # Проверяем больший диапазон ячеек
                                max_rows = min(1000, sheet.max_row)  # Проверяем до 1000 строк
                                max_cols = min(100, sheet.max_column)  # и до 100 столбцов для комментариев
                                
                                for row_idx in range(1, max_rows + 1):
                                    for col_idx in range(1, max_cols + 1):
                                        try:
                                            cell = sheet.cell(row=row_idx, column=col_idx)
                                            if cell.comment is not None:
                                                issues.append("комментарии")
                                                break
                                        except Exception:
                                            continue
                                            
                                    if "комментарии" in issues:
                                        break
                                        
                                if "комментарии" in issues:
                                    break
                            except Exception:
                                continue
                finally:
                    # Закрываем только если wb было инициализировано
                    if wb is not None:
                        wb.close()
        else:
            # Используем win32com для .xls с ограничениями
            pythoncom.CoInitialize()
            excel_app = None
            try:
                excel_app = win32com.client.Dispatch("Excel.Application")
                excel_app.Visible = False
                excel_app.DisplayAlerts = False
                excel_app.ScreenUpdating = False  # Выключаем обновление экрана для ускорения
                
                wb = None
                try:
                    wb = excel_app.Workbooks.Open(file_path, ReadOnly=True, UpdateLinks=False)
                    
                    # Проверка только первых нескольких листов
                    max_sheets = min(3, wb.Sheets.Count)
                    
                    for i in range(1, max_sheets + 1):
                        sheet = wb.Sheets(i)
                        
                        # Проверка на цвет листа
                        try:
                            if sheet.Tab.ColorIndex in [6, 3, 36, 44, 46]:  # Расширенный список: 6-желтый, 3-красный и другие оттенки
                                issues.append(f"цветной лист ({sheet.Name})")
                                break
                        except:
                            pass
                        
                        # Если цвет не определился по ColorIndex, пробуем альтернативный метод через RGB
                        if not "цветной лист" in issues:
                            try:
                                tab_color = sheet.Tab.Color
                                if tab_color:
                                    # Преобразуем в RGB для проверки
                                    r = tab_color % 256
                                    g = (tab_color // 256) % 256
                                    b = (tab_color // 65536) % 256
                                    
                                    # Проверка на желтый (высокие значения R и G, низкое B)
                                    if (r > 200 and g > 200 and b < 100) or (r > 200 and g < 100 and b < 100):
                                        issues.append(f"цветной лист ({sheet.Name})")
                                        break
                            except:
                                pass
                        
                        # Проверка на желтые ячейки - расширяем область проверки
                        if not "желтые ячейки" in issues:
                            try:
                                used_range = sheet.UsedRange
                                if used_range:
                                    # Определяем размеры используемого диапазона с увеличенными пределами
                                    rows_count = min(1500, used_range.Rows.Count)  # Проверяем максимум 1500 строк
                                    cols_count = min(500, used_range.Columns.Count)  # Проверяем максимум 500 столбцов
                                    
                                    # Оптимизация: проверяем сначала небольшой сектор, а затем расширяем
                                    for check_level in range(1, 4):  # 3 уровня проверки с увеличением области
                                        check_rows = min(rows_count, check_level * 500)  # Постепенно увеличиваем до макс. 1500 строк
                                        check_cols = min(cols_count, check_level * 150)  # Постепенно увеличиваем до макс. 500 столбцов
                                        
                                        for row in range(1, check_rows + 1):
                                            for col in range(1, check_cols + 1):
                                                try:
                                                    cell = sheet.Cells(row, col)
                                                    if cell and cell.Interior:
                                                        # Проверка на желтый цвет через ColorIndex (6 - желтый в Excel)
                                                        if cell.Interior.ColorIndex == 6:
                                                            issues.append("желтые ячейки")
                                                            break
                                                        
                                                        # Дополнительная проверка через RGB (для нестандартных цветов)
                                                        try:
                                                            cell_color = cell.Interior.Color
                                                            if cell_color:
                                                                # Извлекаем RGB компоненты
                                                                r = cell_color % 256
                                                                g = (cell_color // 256) % 256
                                                                b = (cell_color // 65536) % 256
                                                                
                                                                # Проверяем похожесть на желтый (высокие R и G, низкий B)
                                                                if r > 200 and g > 200 and b < 100:
                                                                    issues.append("желтые ячейки")
                                                                    break
                                                        except:
                                                            pass
                                                except:
                                                    pass
                                                    
                                            if "желтые ячейки" in issues:
                                                break
                                                
                                        if "желтые ячейки" in issues:
                                            break
                            except:
                                pass
                        
                        # Проверка на комментарии
                        if not "комментарии" in issues:
                            try:
                                if sheet.Comments and sheet.Comments.Count > 0:
                                    issues.append("комментарии")
                            except:
                                pass
                        
                        # Если все проблемы уже найдены, прекращаем проверку
                        if "комментарии" in issues and "желтые ячейки" in issues and "цветной лист" in issues:
                            break
                finally:
                    # Закрываем только если wb было инициализировано
                    if wb is not None:
                        wb.Close(False)
            finally:
                # Выходим из Excel только если приложение было запущено
                if excel_app is not None:
                    excel_app.Quit()
        
        # Поиск заданных пользователем значений
        found_values = []
        if enable_value_search and search_values and isinstance(search_values, list) and len(search_values) > 0:
            try:
                # Для XLSX и XLSM
                if file_path.lower().endswith(('.xlsx', '.xlsm')):
                    # Если мы еще не загрузили workbook ранее, загружаем его сейчас
                    if wb is None:
                        try:
                            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                        except Exception as e:
                            print(f"Не удалось загрузить файл для поиска значений: {str(e)}")
                    
                    if wb is not None:
                        # Проверяем каждый лист
                        for sheet_name in wb.sheetnames:
                            if len(found_values) == len(search_values):
                                break  # Найдены все значения, прекращаем поиск
                                
                            sheet = wb[sheet_name]
                            max_rows = min(2000, sheet.max_row)  # Ограничиваем область поиска
                            max_cols = min(500, sheet.max_column)
                            
                            for row_idx in range(1, max_rows + 1):
                                for col_idx in range(1, max_cols + 1):
                                    try:
                                        cell = sheet.cell(row=row_idx, column=col_idx)
                                        if cell.value:
                                            cell_str_value = str(cell.value)
                                            # Проверяем каждое искомое значение
                                            for value in search_values:
                                                if value in cell_str_value and value not in found_values:
                                                    found_values.append(value)
                                    except Exception:
                                        continue
                                        
                                if len(found_values) == len(search_values):
                                    break  # Найдены все значения, прекращаем поиск
                
                # Для XLS файлов используем win32com
                elif file_path.lower().endswith('.xls') and not found_values:
                    # Если не все значения найдены и мы имеем дело с XLS
                    pythoncom.CoInitialize()
                    excel_app = None
                    try:
                        excel_app = win32com.client.Dispatch("Excel.Application")
                        excel_app.Visible = False
                        excel_app.DisplayAlerts = False
                        excel_app.ScreenUpdating = False
                        
                        wb_search = None
                        try:
                            wb_search = excel_app.Workbooks.Open(file_path, ReadOnly=True, UpdateLinks=False)
                            
                            for i in range(1, wb_search.Sheets.Count + 1):
                                if len(found_values) == len(search_values):
                                    break  # Найдены все значения, прекращаем поиск
                                    
                                sheet = wb_search.Sheets(i)
                                used_range = sheet.UsedRange
                                
                                if used_range:
                                    rows_count = min(2000, used_range.Rows.Count)
                                    cols_count = min(500, used_range.Columns.Count)
                                    
                                    for row in range(1, rows_count + 1):
                                        for col in range(1, cols_count + 1):
                                            try:
                                                cell = sheet.Cells(row, col)
                                                if cell.Value:
                                                    cell_str_value = str(cell.Value)
                                                    for value in search_values:
                                                        if value in cell_str_value and value not in found_values:
                                                            found_values.append(value)
                                            except:
                                                pass
                                                
                                        if len(found_values) == len(search_values):
                                            break  # Найдены все значения
                        finally:
                            if wb_search:
                                wb_search.Close(False)
                    finally:
                        if excel_app:
                            excel_app.Quit()
            except Exception as e:
                print(f"Ошибка при поиске значений: {str(e)}")
                
            # Если найдены значения, добавляем их в список проблем
            if found_values:
                issues.append(f"найдены заданные значения: {', '.join(found_values)}")
                
        # Формирование результата
        if issues:
            result = "Не пройден"
            comment = "Найдено: " + ", ".join(set(issues))
        else:
            result = "Пройден"
            comment = "Проблем не обнаружено"
        
        return {
            'file_name': file_name,
            'file_type': "Excel",
            'file_path': file_path,
            'result': result,
            'comment': comment
        }
        
    except Exception as e:
        return {
            'file_name': file_name,
            'file_type': "Excel",
            'file_path': file_path,
            'result': "Ошибка",
            'comment': f"Ошибка проверки: {str(e)}"
        }